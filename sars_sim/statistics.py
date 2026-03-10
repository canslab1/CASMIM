"""Statistics tracking and file output for the SARS epidemic simulation (ported from C++)."""

import os

import numpy as np

from .models import MeasureData, StateEnum


class StatisticsTracker:
    """Maintains per-day statistics and produces chart-ready data dicts.

    The tracker works with the shared :class:`SimulationData` and
    :class:`MeasureData` objects, computing daily deltas and running
    averages that feed both the live GUI charts and the file output.
    """

    def __init__(self, data, measure):
        """
        Parameters
        ----------
        data : SimulationData
            Mutable simulation state (``statistic``, ``old_statistic``, etc.).
        measure : MeasureData
            Accumulative / average arrays across multiple runs.
        """
        self.data = data
        self.measure = measure

    # ------------------------------------------------------------------
    def backup(self):
        """Copy current ``statistic`` to ``old_statistic`` before the day step."""
        self.data.old_statistic[:] = self.data.statistic[:]

    # ------------------------------------------------------------------
    def update_graphic_data(self):
        """Calculate deltas and update measure arrays.

        Returns
        -------
        dict
            A dictionary keyed by chart name with the values needed to
            refresh the GUI charts:

            * ``'accumulative'`` -- cumulative counts for states 1..6
            * ``'daily'`` -- daily deltas for states 1..6
            * ``'isolated_daily'`` -- daily change in ISOLATED count
            * ``'died_daily'`` -- daily change in DIED count
            * ``'isolated_avg'`` -- running average of ISOLATED + DIED
            * ``'infected_by_normal'`` -- cumulative normal infections
            * ``'infected_by_hospital'`` -- cumulative hospital infections
            * ``'quarantined_total'`` -- cumulative quarantined count
            * ``'quarantined_daily'`` -- daily change in quarantined count
        """
        day = self.data.day
        size = self.measure.size
        stat = self.data.statistic
        old = self.data.old_statistic

        # Accumulate daily deltas into measure arrays for states 1..6
        # (EXPOSED through DIED, plus ISOLATED).
        for i in range(1, StateEnum.SIZE - 1):  # 1 to 6
            delta = int(stat[i] - old[i])
            if day < MeasureData.MAX_DAYS:
                self.measure.value1[day, i] += delta
                self.measure.value2[day, i] = (
                    self.measure.value1[day, i] / float(size) if size > 0 else 0.0
                )

        # Build the chart data dictionary.
        chart_data = {}

        # Cumulative counts: EXPOSED(1)..DIED(5) + ISOLATED(6)
        chart_data['accumulative'] = {
            i: int(stat[i]) for i in range(1, 7)
        }

        # Daily deltas for the same states.
        chart_data['daily'] = {
            i: int(stat[i] - old[i]) for i in range(1, 7)
        }

        chart_data['isolated_daily'] = int(
            stat[StateEnum.ISOLATED] - old[StateEnum.ISOLATED]
        )
        chart_data['died_daily'] = int(
            stat[StateEnum.DIED] - old[StateEnum.DIED]
        )

        if day < MeasureData.MAX_DAYS:
            chart_data['isolated_avg'] = float(
                self.measure.value2[day, StateEnum.ISOLATED]
                + self.measure.value2[day, StateEnum.DIED]
            )
        else:
            chart_data['isolated_avg'] = 0.0

        chart_data['infected_by_normal'] = self.data.infected_by_normal
        chart_data['infected_by_hospital'] = self.data.infected_by_hospital
        chart_data['quarantined_total'] = int(stat[StateEnum.QUARANTINED])
        chart_data['quarantined_daily'] = int(
            stat[StateEnum.QUARANTINED] - old[StateEnum.QUARANTINED]
        )

        return chart_data

    # ------------------------------------------------------------------
    def get_state_counts(self):
        """Count people in each state by scanning ``people_state``.

        Uses ``np.bincount`` for a single-pass O(N) scan instead of
        8 separate full-array comparisons.

        Returns
        -------
        dict[int, int]
            Mapping from ``StateEnum`` value to the number of people in
            that state.
        """
        N = self.data.N
        states = self.data.people_state[:N]
        bins = np.bincount(states.astype(np.intp), minlength=StateEnum.SIZE)
        return {s: int(bins[s]) for s in range(StateEnum.SIZE)}


# ======================================================================
# File output (Excel)
# ======================================================================

class FileOutput:
    """Produces a single Excel workbook with four sheets for simulation output.

    Sheet layout (mirrors the original C++ ``outFile1`` .. ``outFile4``):

    * **Output 1** -- Cumulative state counts + mortality rate per day.
    * **Output 2** -- Daily new + cumulative counts for all 8 states.
    * **Output 3** -- Free-form action / event log.
    * **Output 4** -- Daily running-average counts for all 8 states.
    """

    _STATE_NAMES = [
        'Susceptible', 'Exposed', 'Infective', 'Recovered',
        'Immune', 'Died', 'Isolated', 'Quarantined',
    ]

    def __init__(self):
        self.wb = None          # openpyxl Workbook
        self.sheets = [None] * 4
        self.enabled = [False] * 4
        self.directory = None
        self._filepath = None
        self._action_row = 2    # next row for Output 3

    # ------------------------------------------------------------------
    def init_files(self, serial, enabled_flags):
        """Create the output directory and initialise the Excel workbook.

        Parameters
        ----------
        serial : int or str
            Timestamp (or other unique token) used as the directory name.
        enabled_flags : list[bool]
            Four booleans controlling which sheets receive data.

        Returns
        -------
        bool
            ``True`` on success, ``False`` if the directory cannot be created.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment

        self.directory = str(serial)
        self.enabled = list(enabled_flags)

        try:
            os.makedirs(self.directory, exist_ok=False)
        except OSError:
            return False

        self._filepath = os.path.join(self.directory, 'OUTPUT.xlsx')

        self.wb = Workbook()
        # Remove default sheet created by openpyxl
        self.wb.remove(self.wb.active)

        bold = Font(bold=True)
        center = Alignment(horizontal='center')

        # ---- Sheet 1: cumulative counts + mortality ----
        if self.enabled[0]:
            ws = self.wb.create_sheet('Output 1')
            self.sheets[0] = ws
            headers = [
                'Day', 'Susceptible', 'Exposed', 'Infective',
                'Recovered', 'Immune', 'Died',
                'Cum. Recovered', 'Cum. Died', 'Mortality %',
            ]
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=col, value=h)
                c.font = bold
                c.alignment = center

        # ---- Sheet 2: daily new + cumulative ----
        if self.enabled[1]:
            ws = self.wb.create_sheet('Output 2')
            self.sheets[1] = ws
            headers = ['Day']
            for prefix in ('Daily ', 'Cum. '):
                headers += [prefix + n for n in self._STATE_NAMES]
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=col, value=h)
                c.font = bold
                c.alignment = center

        # ---- Sheet 3: action log ----
        if self.enabled[2]:
            ws = self.wb.create_sheet('Output 3')
            self.sheets[2] = ws
            c = ws.cell(row=1, column=1, value='Action Log')
            c.font = bold
            self._action_row = 2

        # ---- Sheet 4: daily running averages ----
        if self.enabled[3]:
            ws = self.wb.create_sheet('Output 4')
            self.sheets[3] = ws
            headers = ['Day'] + ['Avg. ' + n for n in self._STATE_NAMES]
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=col, value=h)
                c.font = bold
                c.alignment = center

        return True

    # ------------------------------------------------------------------
    def write_daily(self, day, statistic, old_statistic, state_counts, measure_data):
        """Write one day's row to each enabled sheet.

        Parameters
        ----------
        day : int
            Current simulation day (0-based).
        statistic : numpy.ndarray
            Current cumulative statistic array (length ``StateEnum.SIZE``).
        old_statistic : numpy.ndarray
            Previous day's cumulative statistic array.
        state_counts : dict[int, int]
            Per-state population counts.
        measure_data : MeasureData
            Running-average data for sheet 4.
        """
        stat = statistic
        old = old_statistic

        # ---- Sheet 1: cumulative counts + mortality ----
        if self.sheets[0] and self.enabled[0]:
            sc = state_counts
            total_resolved = stat[StateEnum.RECOVERED] + stat[StateEnum.DIED]
            mortality = (
                (stat[StateEnum.DIED] / total_resolved * 100.0)
                if total_resolved > 0 else 0.0
            )
            row = [
                day,
                sc.get(0, 0), sc.get(1, 0), sc.get(2, 0),
                sc.get(3, 0), sc.get(4, 0), sc.get(5, 0),
                int(stat[StateEnum.RECOVERED]),
                int(stat[StateEnum.DIED]),
                round(mortality, 2),
            ]
            self.sheets[0].append(row)

        # ---- Sheet 2: daily new + cumulative for all 8 states ----
        if self.sheets[1] and self.enabled[1]:
            row = [day]
            for i in range(StateEnum.SIZE):
                row.append(int(stat[i] - old[i]))
            for i in range(StateEnum.SIZE):
                row.append(int(stat[i]))
            self.sheets[1].append(row)

        # ---- Sheet 4: daily running averages for all 8 states ----
        if self.sheets[3] and self.enabled[3]:
            row = [day]
            for i in range(StateEnum.SIZE):
                val = (
                    float(measure_data.value2[day, i])
                    if day < MeasureData.MAX_DAYS else 0.0
                )
                row.append(round(val, 2))
            self.sheets[3].append(row)

    # ------------------------------------------------------------------
    def write_action(self, msg):
        """Append a free-form action / event message to Output 3.

        Parameters
        ----------
        msg : str
            One line of text to log.
        """
        if self.sheets[2] and self.enabled[2]:
            self.sheets[2].cell(row=self._action_row, column=1, value=msg)
            self._action_row += 1

    # ------------------------------------------------------------------
    def close(self, measure_size=0):
        """Save and close the Excel workbook.

        Parameters
        ----------
        measure_size : int, optional
            Number of simulation runs to report in the Output 4 footer.
        """
        if self.wb is None:
            return

        # Footer for Output 4
        if self.sheets[3] and self.enabled[3]:
            self.sheets[3].append([])
            self.sheets[3].append([f'上述數據為 {measure_size} 次模擬的結果'])

        # Auto-adjust column widths for readability
        for ws in self.wb.worksheets:
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_len + 3, 20)

        try:
            self.wb.save(self._filepath)
        except OSError:
            pass  # silently ignore save failure (e.g. disk full, permission)
        finally:
            self.wb.close()
            self.wb = None
            self.sheets = [None] * 4
